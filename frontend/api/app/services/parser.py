"""Data parser service: reads CSV/XLSX files and produces structured summaries without heavy dependencies."""

import io
import csv
from openpyxl import load_workbook


def parse_file(contents: bytes, filename: str) -> dict:
    """Parse uploaded file into a structured data summary without Pandas.

    Args:
        contents: Raw file bytes.
        filename: Original filename (used for extension detection).

    Returns:
        Dictionary with parsed data summary.
    """
    ext = filename.rsplit(".", 1)[-1].lower()
    
    rows = []
    headers = []

    if ext == "csv":
        stream = io.StringIO(contents.decode("utf-8", errors="ignore"))
        reader = csv.DictReader(stream)
        headers = reader.fieldnames or []
        rows = list(reader)
    elif ext == "xlsx":
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        sheet = wb.active
        # Convert sheet to list of dicts
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell is not None else f"Col_{j}" for j, cell in enumerate(row)]
            else:
                rows.append(dict(zip(headers, row)))
    else:
        raise ValueError(f"Unsupported file extension: .{ext}")

    # Build a comprehensive summary for the AI
    summary = {
        "filename": filename,
        "total_rows": len(rows),
        "total_columns": len(headers),
        "columns": headers,
        "sample_rows": rows[:10],
        "statistics": {},
    }

    # Basic statistics and revenue aggregation
    total_revenue = 0.0
    total_units_sold = 0
    has_revenue = "Revenue" in headers
    has_units = "Units_Sold" in headers

    for row in rows:
        if has_revenue:
            try:
                val = float(str(row.get("Revenue", 0)).replace("$", "").replace(",", ""))
                total_revenue += val
            except (ValueError, TypeError):
                pass
        if has_units:
            try:
                val = int(float(str(row.get("Units_Sold", 0)).replace(",", "")))
                total_units_sold += val
            except (ValueError, TypeError):
                pass

    if has_revenue:
        summary["total_revenue"] = round(total_revenue, 2)
    if has_units:
        summary["total_units_sold"] = total_units_sold

    return summary
